from django.http import HttpResponse
from django.contrib.auth import get_user_model
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import (
    AllowAny,
    IsAuthenticated
)
from rest_framework import mixins, status, viewsets
# from apps_com_teedii_core.users.serializers import UserSerializer
from django_filters.rest_framework import DjangoFilterBackend

from Evote.models import VoteStudent, Period

from ..serializers import (
    UserSerializer,
    UserLoginSerializer

)

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

# models

User = get_user_model()


class UserViewSet(viewsets.ModelViewSet):

    serializer_class = UserSerializer
    queryset = User.objects.all()

    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('identificacion',)

    def get_permissions(self):
        """Assign permissions based on action."""
        if self.action in ['login', ]:
            permissions = [AllowAny]
        elif self.action in ['retrieve', 'update', 'partial_update', ]:
            permissions = [IsAuthenticated]
        else:
            permissions = [IsAuthenticated]
        return [p() for p in permissions]

    @action(detail=False, methods=['post'])
    def login(self, request):
        """User sign in."""

        serializer = UserLoginSerializer(data=request.data)

        is_valid = serializer.is_valid()

        # un jemplo se puede mejorar esto
        # https://stackoverflow.com/questions/29731013/django-rest-framework-cannot-call-is-valid-as-no-data-keyword-argument/29731923
        if(is_valid):

            user, token = serializer.save()
            user = UserSerializer(user).data

            data = {
                'user': user,
                'access_token': token
            }

        else:
            data = {
                'error': '001',
                'message': 'Usuario o contraseña incorrecto'
            }

        return Response(data, status=status.HTTP_201_CREATED)


def report(request):
    students = User.objects.all()
    period = Period.objects.get(periodPresent=True)
    list_students = []
    for student in students:
        stu = {}
        if student.carrera != None:
            stu['id'] = student.identificacion
            stu['nombres'] = student.nombres
            stu['apellidos'] = student.apellidos
            stu['carrera'] = student.carrera
            stu['ciclo'] = student.ciclo
            stu['paralelo'] = student.paralelo

            try:
                VoteStudent.objects.get(period=period, student=student)
                stu['voto'] = "SI"
            except VoteStudent.DoesNotExist:
                stu['voto'] = "NO"

            list_students.append(stu)

    wb = Workbook()
    sheet = wb.active

    c1 = sheet.cell(row=1, column=1)
    c1.value = "IDENTIFICACIÓ"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "NOMBRES"
    c3 = sheet.cell(row=1, column=3)
    c3.value = "APELLIDOS"
    c4 = sheet.cell(row=1, column=4)
    c4.value = "CARRERA"
    c5 = sheet.cell(row=1, column=5)
    c5.value = "CICLO"
    c6 = sheet.cell(row=1, column=6)
    c6.value = "PARALELO"
    c7 = sheet.cell(row=1, column=7)
    c7.value = "VOTO"

    # export data to Excel
    rows = User.objects.all().values_list('identificacion', 'nombres',
                                          'apellidos', 'carrera', 'ciclo', 'paralelo',)

    rows = rows.exclude(carrera=None)
    for row_num, row in enumerate(rows, 1):
        for col_num, value in enumerate(row):
            c8 = sheet.cell(row=row_num+1, column=col_num+1)
            c8.value = value
            try:
                VoteStudent.objects.get(period=period, student=row)
                c9 = sheet.cell(row=row_num+1, column=col_num+2)
                c9.value = "SI"
            except VoteStudent.DoesNotExist:
                c9 = sheet.cell(row=row_num+1, column=col_num+2)
                c9.value = "NO"

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=votos_etudiantes.xlsx'
    wb.save(response)

    return response
